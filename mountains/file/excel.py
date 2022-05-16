# -*- coding: utf-8 -*-
# created by restran on 2018/04/16
from __future__ import unicode_literals, absolute_import

from datetime import datetime

from ..base import BytesIO, text_type, iteritems
from ..file import write_bytes_file

__all__ = ['read_excel', 'write_excel', 'edit_excel']


def read_excel(file_name=None, file_contents=None, offset=1,
               header_index=0, sheet_index=0, sheet_name=None,
               dt2str=True, lower_header=True):
    """
    读取 Excel
    :param sheet_name:
    :param dt2str: 将日期类型的数据转成字符串
    :param header_index: header 在哪一行
    :param file_contents:
    :param sheet_index:
    :param file_name:
    :param offset: 偏移，一般第一行是表头，不需要读取数据
    :param lower_header: 标题转成小写
    :return:
    """
    try:
        import xlrd
        from xlrd import xldate_as_tuple
    except:
        raise Exception('xlrd is not installed')

    try:
        workbook = xlrd.open_workbook(filename=file_name, file_contents=file_contents)
    except Exception as e:
        return None

    if len(workbook.sheets()) <= 0:
        return []

    if sheet_name is not None:
        sh = workbook.sheet_by_name(sheet_name)
    else:
        sh = workbook.sheet_by_index(sheet_index)

    raw_data = []
    n_rows = sh.nrows
    row = sh.row_values(header_index)
    header = []
    for t in row:
        t = text_type(t).strip()
        if lower_header:
            t = t.lower()
        header.append(t)

    # n_cols = sh.ncols
    # 第0行是提示信息和标题，跳过
    for i in range(offset, n_rows):
        try:
            # row = sh.row_values(i)
            d = {}
            # ctype: 0 empty,1 string, 2 number, 3 date, 4 boolean, 5 error
            for j, t in enumerate(header):
                ctype = sh.cell(i, j).ctype  # 表格的数据类型
                cell = sh.cell_value(i, j)
                if ctype == 2 and cell % 1 == 0:  # 如果是整形
                    cell = int(cell)
                elif ctype == 3:
                    # 转成datetime对象
                    cell = datetime(*xldate_as_tuple(cell, 0))
                    if dt2str:
                        cell = cell.strftime('%Y-%m-%d %H:%M:%S')
                elif ctype == 4:
                    cell = True if cell == 1 else False

                d[t] = cell
            raw_data.append(d)
        except Exception as e:
            pass

    return raw_data


def write_excel(headers, data, file_name, file_io=None):
    """
    写数据到新的Excel中
    :param headers:
    :param data:
    :param file_name:
    :param file_io:
    :return:
    """
    try:
        import xlsxwriter
    except:
        raise Exception('xlsxwriter is not installed')

    sio = BytesIO()
    workbook = xlsxwriter.Workbook(sio)
    worksheet = workbook.add_worksheet()
    new_headers = []
    if len(headers) > 0:
        for t in headers:
            if isinstance(t, dict):
                new_headers.append(t.get('name', ''))
            else:
                new_headers.append(t)
    headers = new_headers
    for i, t in enumerate(headers):
        worksheet.write(0, i, t)

    index = 1
    for row in data:
        if isinstance(row, dict):
            for i, name in enumerate(headers):
                worksheet.write(index, i, row.get(name, ''))
        else:
            for i, x in enumerate(row):
                worksheet.write(index, i, x)
        index += 1
    # 关闭 Excel
    workbook.close()
    if file_io is not None:
        if not isinstance(file_io, BytesIO):
            raise Exception('output_fio should be BytesIO')
        else:
            file_io.write(sio.getvalue())
    else:
        write_bytes_file(file_name, sio.getvalue())


def edit_excel(file_name=None, sheet_index=0, sheet_name=None, data=None, output_filename=None, output_fio=None):
    """
    编辑 Excel，打开已有的 Excel，往里面填充数据
    :param file_name:
    :param sheet_index:
    :param sheet_name:
    :param data: data = {'A2': '123', 'A3': '456'}
    :param output_filename:
    :param output_fio:
    :type data: dict
    """
    try:
        from openpyxl import load_workbook
    except:
        raise Exception('openpyxl is not installed')

    try:
        wb = load_workbook(file_name)
        if sheet_name is not None:
            ws = wb[sheet_name]
        else:
            ws = wb.worksheets[sheet_index]
    except Exception as e:
        return None

    for key, value in iteritems(data):
        try:
            ws[key] = value
        except:
            pass

    if output_fio is not None:
        if not isinstance(output_fio, BytesIO):
            raise Exception('output_fio should be BytesIO')
        else:
            wb.save(output_fio)
    elif output_filename is not None:
        wb.save(output_filename)
    else:
        wb.save(file_name)

    wb.close()


def to_excel_column_no(index, index_start_1=False):
    """
    Excel 列编号是: A, B, ..., AA, ..., BA
    index 编号默认从1开始
    """
    if index_start_1:
        index -= 1
    y = ''
    s = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'

    while index >= 0:
        y += s[index % 26]
        index = index // 26 - 1

    return y[::-1]
